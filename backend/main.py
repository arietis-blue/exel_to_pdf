from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os

app = FastAPI()

# CORS設定の追加
origins = [
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/process/")
async def process_file(file: UploadFile = File(...)):
    contents = await file.read()
    with open(file.filename, 'wb') as f:
        f.write(contents)

    # 特定の範囲のセルを抽出
    workbook = openpyxl.load_workbook(file.filename)
    sheet = workbook.active
    extracted_data = []
    for row in sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=4, values_only=True):
        extracted_data.append(row)

    # 新しいエクセルファイルにデータをまとめる
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    for row_idx, row_data in enumerate(extracted_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    new_excel_file = "new_file.xlsx"
    new_workbook.save(new_excel_file)

    # PDFに変換
    pdf_output = "output.pdf"
    c = canvas.Canvas(pdf_output, pagesize=A4)
    width, height = A4

    row_height = 20  # 一行の高さ
    y = height - row_height  # 初期Y位置

    for row_data in extracted_data:
        x = 40  # 初期X位置
        for cell_value in row_data:
            c.drawString(x, y, str(cell_value) if cell_value is not None else "")
            x += 100  # 列の幅
        y -= row_height
        if y < row_height:  # 新しいページを追加
            c.showPage()
            y = height - row_height

    c.save()

    # 元のファイルを削除
    os.remove(file.filename)

    return FileResponse(pdf_output, media_type='application/pdf')

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
